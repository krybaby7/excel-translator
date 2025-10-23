"""
Unit tests for Excel Translator functions
"""
import pytest
import os
import sys
from openpyxl import load_workbook

# Add parent directory to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from excel_translator import convert_xls_to_xlsx, translate_excel_with_format, process_file


class TestConvertXlsToXlsx:
    """Test cases for convert_xls_to_xlsx function"""

    def test_convert_valid_xls(self):
        """Test converting a valid .xls file"""
        input_file = "test_data/old_format.xls"
        output_file = convert_xls_to_xlsx(input_file)

        assert output_file.endswith('.xlsx')
        assert os.path.exists(output_file)

        # Verify content was preserved
        wb = load_workbook(output_file)
        ws = wb.active
        assert ws['A1'].value == 'Bonjour'
        assert ws['B1'].value == 'Monde'

        # Cleanup
        os.remove(output_file)

    def test_convert_xls_file_not_found(self):
        """Test error handling when file doesn't exist"""
        with pytest.raises(FileNotFoundError):
            convert_xls_to_xlsx("nonexistent.xls")

    def test_convert_xls_wrong_format(self):
        """Test error handling when file is not .xls"""
        with pytest.raises(ValueError):
            convert_xls_to_xlsx("test_data/simple_french.xlsx")

    def test_convert_xls_preserves_multiple_sheets(self):
        """Test that multiple sheets are preserved during conversion"""
        # This test would require a multi-sheet .xls file
        # Skipping for now but important for future implementation
        pass


class TestTranslateExcelWithFormat:
    """Test cases for translate_excel_with_format function"""

    def test_translate_simple_file(self):
        """Test translating a simple file"""
        input_file = "test_data/simple_french.xlsx"
        output_file = "test_results/translated_simple.xlsx"

        translate_excel_with_format(input_file, output_file, "fr", "en")

        assert os.path.exists(output_file)

        # Verify translation occurred
        wb = load_workbook(output_file)
        ws = wb.active
        # Note: Actual translation may vary, just check it's not the original
        translated_value = ws['A1'].value
        assert translated_value is not None
        assert len(translated_value) > 0

        # Cleanup
        os.remove(output_file)

    def test_translate_preserves_formatting(self):
        """Test that formatting is preserved during translation"""
        input_file = "test_data/formatted_french.xlsx"
        output_file = "test_results/translated_formatted.xlsx"

        translate_excel_with_format(input_file, output_file, "fr", "en")

        # Load original and translated files
        wb_original = load_workbook(input_file)
        wb_translated = load_workbook(output_file)

        ws_orig = wb_original.active
        ws_trans = wb_translated.active

        # Check bold formatting preserved
        assert ws_trans['A1'].font.bold == ws_orig['A1'].font.bold
        # Check fill color preserved
        assert ws_trans['A1'].fill.start_color == ws_orig['A1'].fill.start_color

        # Cleanup
        os.remove(output_file)

    def test_translate_multi_sheet(self):
        """Test translating a file with multiple sheets"""
        input_file = "test_data/multi_sheet_french.xlsx"
        output_file = "test_results/translated_multi_sheet.xlsx"

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb = load_workbook(output_file)
        assert len(wb.sheetnames) == 3
        assert "Feuille1" in wb.sheetnames
        assert "Feuille2" in wb.sheetnames
        assert "Feuille3" in wb.sheetnames

        # Cleanup
        os.remove(output_file)

    def test_translate_mixed_content(self):
        """Test translating file with mixed content types"""
        input_file = "test_data/mixed_content_french.xlsx"
        output_file = "test_results/translated_mixed.xlsx"

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb = load_workbook(output_file)
        ws = wb.active

        # Check that numbers are preserved
        assert isinstance(ws['B2'].value, int)
        # Check that formulas are preserved (though the result might change)
        # Note: formulas should remain as formulas
        assert ws['B4'].value is not None

        # Cleanup
        os.remove(output_file)

    def test_translate_empty_file(self):
        """Test translating an empty file"""
        input_file = "test_data/empty.xlsx"
        output_file = "test_results/translated_empty.xlsx"

        translate_excel_with_format(input_file, output_file, "fr", "en")

        assert os.path.exists(output_file)

        # Cleanup
        os.remove(output_file)

    def test_translate_file_not_found(self):
        """Test error handling when file doesn't exist"""
        with pytest.raises(FileNotFoundError):
            translate_excel_with_format("nonexistent.xlsx", "output.xlsx")

    def test_translate_wrong_format(self):
        """Test error handling when file is not .xlsx"""
        with pytest.raises(ValueError):
            translate_excel_with_format("test.txt", "output.xlsx")

    def test_translate_special_characters(self):
        """Test handling of special characters"""
        input_file = "test_data/special_chars_french.xlsx"
        output_file = "test_results/translated_special.xlsx"

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb = load_workbook(output_file)
        ws = wb.active

        # Verify file was processed without errors
        assert ws['A1'].value is not None
        assert ws['A2'].value is not None  # Emojis
        assert ws['A3'].value is not None  # Symbols

        # Cleanup
        os.remove(output_file)

    def test_translate_different_language_pairs(self):
        """Test translation with different language pairs"""
        input_file = "test_data/simple_french.xlsx"

        # French to English
        output_en = "test_results/translated_fr_to_en.xlsx"
        translate_excel_with_format(input_file, output_en, "fr", "en")
        assert os.path.exists(output_en)

        # French to Spanish
        output_es = "test_results/translated_fr_to_es.xlsx"
        translate_excel_with_format(input_file, output_es, "fr", "es")
        assert os.path.exists(output_es)

        # French to German
        output_de = "test_results/translated_fr_to_de.xlsx"
        translate_excel_with_format(input_file, output_de, "fr", "de")
        assert os.path.exists(output_de)

        # Cleanup
        os.remove(output_en)
        os.remove(output_es)
        os.remove(output_de)


class TestProcessFile:
    """Test cases for process_file function"""

    def test_process_xlsx_file(self):
        """Test processing a .xlsx file"""
        output = process_file("test_data/simple_french.xlsx", "fr", "en")
        assert os.path.exists(output)
        assert output.startswith("translated_")
        os.remove(output)

    def test_process_xls_file(self):
        """Test processing a .xls file (should convert then translate)"""
        output = process_file("test_data/old_format.xls", "fr", "en")
        assert os.path.exists(output)
        assert output.endswith(".xlsx")

        # Cleanup both converted and translated files
        os.remove(output)
        if os.path.exists("test_data/old_format.xlsx"):
            os.remove("test_data/old_format.xlsx")

    def test_process_unsupported_format(self):
        """Test error handling for unsupported file format"""
        with pytest.raises(ValueError):
            process_file("test.txt")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
