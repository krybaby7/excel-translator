"""
Data integrity validation tests
Ensures no data is lost or corrupted during translation
"""
import pytest
import os
import sys
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from excel_translator import translate_excel_with_format, convert_xls_to_xlsx


class TestDataIntegrity:
    """Test that data integrity is maintained"""

    def test_cell_count_preserved(self):
        """Verify the same number of cells before and after"""
        input_file = "test_data/simple_french.xlsx"
        output_file = "test_results/integrity_cell_count.xlsx"

        # Count cells in original
        wb_orig = load_workbook(input_file)
        ws_orig = wb_orig.active
        orig_cell_count = sum(1 for row in ws_orig.iter_rows() for cell in row if cell.value is not None)

        # Translate
        translate_excel_with_format(input_file, output_file, "fr", "en")

        # Count cells in translated
        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active
        trans_cell_count = sum(1 for row in ws_trans.iter_rows() for cell in row if cell.value is not None)

        assert orig_cell_count == trans_cell_count, \
            f"Cell count mismatch: {orig_cell_count} -> {trans_cell_count}"

        # Cleanup
        os.remove(output_file)

    def test_sheet_count_preserved(self):
        """Verify all sheets are preserved"""
        input_file = "test_data/multi_sheet_french.xlsx"
        output_file = "test_results/integrity_sheet_count.xlsx"

        wb_orig = load_workbook(input_file)
        orig_sheet_count = len(wb_orig.sheetnames)

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        trans_sheet_count = len(wb_trans.sheetnames)

        assert orig_sheet_count == trans_sheet_count, \
            f"Sheet count mismatch: {orig_sheet_count} -> {trans_sheet_count}"

        # Cleanup
        os.remove(output_file)

    def test_sheet_names_preserved(self):
        """Verify sheet names are preserved"""
        input_file = "test_data/multi_sheet_french.xlsx"
        output_file = "test_results/integrity_sheet_names.xlsx"

        wb_orig = load_workbook(input_file)
        orig_sheet_names = wb_orig.sheetnames

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        trans_sheet_names = wb_trans.sheetnames

        assert orig_sheet_names == trans_sheet_names, \
            f"Sheet names changed: {orig_sheet_names} -> {trans_sheet_names}"

        # Cleanup
        os.remove(output_file)

    def test_numeric_values_unchanged(self):
        """Verify numeric values remain unchanged"""
        input_file = "test_data/mixed_content_french.xlsx"
        output_file = "test_results/integrity_numbers.xlsx"

        wb_orig = load_workbook(input_file)
        ws_orig = wb_orig.active
        orig_b2 = ws_orig['B2'].value

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active
        trans_b2 = ws_trans['B2'].value

        assert orig_b2 == trans_b2, f"Numeric value changed: {orig_b2} -> {trans_b2}"

        # Cleanup
        os.remove(output_file)

    def test_empty_cells_remain_empty(self):
        """Verify empty cells stay empty"""
        # Create test file with specific pattern
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Bonjour"
        ws['A2'] = None
        ws['A3'] = "Merci"

        input_file = "test_data/integrity_empty.xlsx"
        wb.save(input_file)

        output_file = "test_results/integrity_empty.xlsx"
        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active

        assert ws_trans['A2'].value is None, "Empty cell was filled"

        # Cleanup
        os.remove(input_file)
        os.remove(output_file)

    def test_font_formatting_preserved(self):
        """Verify font formatting is preserved"""
        input_file = "test_data/formatted_french.xlsx"
        output_file = "test_results/integrity_fonts.xlsx"

        wb_orig = load_workbook(input_file)
        ws_orig = wb_orig.active
        orig_bold = ws_orig['A1'].font.bold
        orig_italic = ws_orig['A2'].font.italic
        orig_underline = ws_orig['A3'].font.underline

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active
        trans_bold = ws_trans['A1'].font.bold
        trans_italic = ws_trans['A2'].font.italic
        trans_underline = ws_trans['A3'].font.underline

        assert orig_bold == trans_bold, "Bold formatting lost"
        assert orig_italic == trans_italic, "Italic formatting lost"
        assert orig_underline == trans_underline, "Underline formatting lost"

        # Cleanup
        os.remove(output_file)

    def test_cell_colors_preserved(self):
        """Verify cell background colors are preserved"""
        input_file = "test_data/formatted_french.xlsx"
        output_file = "test_results/integrity_colors.xlsx"

        wb_orig = load_workbook(input_file)
        ws_orig = wb_orig.active
        orig_fill = ws_orig['A1'].fill.start_color.rgb

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active
        trans_fill = ws_trans['A1'].fill.start_color.rgb

        assert orig_fill == trans_fill, f"Fill color changed: {orig_fill} -> {trans_fill}"

        # Cleanup
        os.remove(output_file)

    def test_alignment_preserved(self):
        """Verify cell alignment is preserved"""
        input_file = "test_data/formatted_french.xlsx"
        output_file = "test_results/integrity_alignment.xlsx"

        wb_orig = load_workbook(input_file)
        ws_orig = wb_orig.active
        orig_horizontal = ws_orig['B1'].alignment.horizontal
        orig_vertical = ws_orig['B1'].alignment.vertical

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active
        trans_horizontal = ws_trans['B1'].alignment.horizontal
        trans_vertical = ws_trans['B1'].alignment.vertical

        assert orig_horizontal == trans_horizontal, "Horizontal alignment changed"
        assert orig_vertical == trans_vertical, "Vertical alignment changed"

        # Cleanup
        os.remove(output_file)

    def test_borders_preserved(self):
        """Verify cell borders are preserved"""
        input_file = "test_data/formatted_french.xlsx"
        output_file = "test_results/integrity_borders.xlsx"

        wb_orig = load_workbook(input_file)
        ws_orig = wb_orig.active
        orig_border_left = ws_orig['C1'].border.left.style
        orig_border_right = ws_orig['C1'].border.right.style

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active
        trans_border_left = ws_trans['C1'].border.left.style
        trans_border_right = ws_trans['C1'].border.right.style

        assert orig_border_left == trans_border_left, "Left border changed"
        assert orig_border_right == trans_border_right, "Right border changed"

        # Cleanup
        os.remove(output_file)

    def test_xls_to_xlsx_data_integrity(self):
        """Verify no data loss during .xls to .xlsx conversion"""
        input_file = "test_data/old_format.xls"

        # Get original data using xlrd
        import xlrd
        wb_xls = xlrd.open_workbook(input_file)
        sheet_xls = wb_xls.sheet_by_index(0)
        orig_value_a1 = sheet_xls.cell_value(0, 0)
        orig_value_b1 = sheet_xls.cell_value(0, 1)

        # Convert
        output_file = convert_xls_to_xlsx(input_file)

        # Check converted data
        wb_xlsx = load_workbook(output_file)
        ws_xlsx = wb_xlsx.active
        new_value_a1 = ws_xlsx['A1'].value
        new_value_b1 = ws_xlsx['B1'].value

        assert orig_value_a1 == new_value_a1, f"A1 value changed: {orig_value_a1} -> {new_value_a1}"
        assert orig_value_b1 == new_value_b1, f"B1 value changed: {orig_value_b1} -> {new_value_b1}"

        # Cleanup
        os.remove(output_file)

    def test_row_count_preserved(self):
        """Verify row count remains the same"""
        input_file = "test_data/large_french.xlsx"
        output_file = "test_results/integrity_row_count.xlsx"

        wb_orig = load_workbook(input_file)
        ws_orig = wb_orig.active
        orig_row_count = ws_orig.max_row

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active
        trans_row_count = ws_trans.max_row

        assert orig_row_count == trans_row_count, \
            f"Row count changed: {orig_row_count} -> {trans_row_count}"

        # Cleanup
        os.remove(output_file)

    def test_column_count_preserved(self):
        """Verify column count remains the same"""
        input_file = "test_data/large_french.xlsx"
        output_file = "test_results/integrity_column_count.xlsx"

        wb_orig = load_workbook(input_file)
        ws_orig = wb_orig.active
        orig_col_count = ws_orig.max_column

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb_trans = load_workbook(output_file)
        ws_trans = wb_trans.active
        trans_col_count = ws_trans.max_column

        assert orig_col_count == trans_col_count, \
            f"Column count changed: {orig_col_count} -> {trans_col_count}"

        # Cleanup
        os.remove(output_file)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
