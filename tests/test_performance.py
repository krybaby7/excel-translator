"""
Performance and benchmark tests for Excel Translator
"""
import pytest
import os
import sys
import time
from openpyxl import Workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from excel_translator import translate_excel_with_format, convert_xls_to_xlsx


class TestPerformance:
    """Performance benchmarks"""

    def test_small_file_performance(self, benchmark):
        """Benchmark translation of small file (10 cells)"""
        input_file = "test_data/simple_french.xlsx"
        output_file = "test_results/perf_small.xlsx"

        def translate_small():
            translate_excel_with_format(input_file, output_file, "fr", "en")

        result = benchmark(translate_small)

        # Cleanup
        if os.path.exists(output_file):
            os.remove(output_file)

    def test_medium_file_performance(self, benchmark):
        """Benchmark translation of medium file (100 cells)"""
        # Create medium test file
        wb = Workbook()
        ws = wb.active

        french_phrases = [
            "Bonjour", "Merci", "Au revoir", "Comment allez-vous?",
            "Bonne journée", "S'il vous plaît", "Excusez-moi", "Oui", "Non", "Peut-être"
        ]

        for row in range(1, 11):
            for col in range(1, 11):
                phrase = french_phrases[(row + col) % len(french_phrases)]
                ws.cell(row=row, column=col, value=phrase)

        input_file = "test_data/perf_medium.xlsx"
        wb.save(input_file)

        output_file = "test_results/perf_medium.xlsx"

        def translate_medium():
            translate_excel_with_format(input_file, output_file, "fr", "en")

        result = benchmark(translate_medium)

        # Cleanup
        os.remove(input_file)
        if os.path.exists(output_file):
            os.remove(output_file)

    def test_large_file_performance_timed(self):
        """Time translation of large file (1000x10 = 10,000 cells)"""
        input_file = "test_data/large_french.xlsx"
        output_file = "test_results/perf_large.xlsx"

        start_time = time.time()
        translate_excel_with_format(input_file, output_file, "fr", "en")
        end_time = time.time()

        duration = end_time - start_time
        print(f"\nLarge file translation took: {duration:.2f} seconds")

        # Assert it completes within reasonable time (e.g., 5 minutes)
        assert duration < 300, f"Translation took too long: {duration:.2f}s"

        # Cleanup
        os.remove(output_file)

    def test_multi_sheet_performance(self, benchmark):
        """Benchmark multi-sheet file translation"""
        input_file = "test_data/multi_sheet_french.xlsx"
        output_file = "test_results/perf_multi_sheet.xlsx"

        def translate_multi():
            translate_excel_with_format(input_file, output_file, "fr", "en")

        result = benchmark(translate_multi)

        # Cleanup
        if os.path.exists(output_file):
            os.remove(output_file)

    def test_xls_conversion_performance(self, benchmark):
        """Benchmark .xls to .xlsx conversion"""
        input_file = "test_data/old_format.xls"

        def convert_xls():
            output = convert_xls_to_xlsx(input_file)
            return output

        result = benchmark(convert_xls)

        # Cleanup converted file
        xlsx_file = input_file.replace('.xls', '.xlsx')
        if os.path.exists(xlsx_file):
            os.remove(xlsx_file)

    def test_batch_processing_performance(self):
        """Test processing multiple files in sequence"""
        files = [
            "test_data/simple_french.xlsx",
            "test_data/formatted_french.xlsx",
            "test_data/single_cell.xlsx",
        ]

        start_time = time.time()

        for i, input_file in enumerate(files):
            output_file = f"test_results/perf_batch_{i}.xlsx"
            translate_excel_with_format(input_file, output_file, "fr", "en")

        end_time = time.time()
        duration = end_time - start_time

        print(f"\nBatch processing of {len(files)} files took: {duration:.2f} seconds")
        print(f"Average time per file: {duration/len(files):.2f} seconds")

        # Cleanup
        for i in range(len(files)):
            output_file = f"test_results/perf_batch_{i}.xlsx"
            if os.path.exists(output_file):
                os.remove(output_file)


class TestMemoryUsage:
    """Test memory handling with large files"""

    def test_large_file_memory(self):
        """Test that large file doesn't cause memory issues"""
        input_file = "test_data/large_french.xlsx"
        output_file = "test_results/memory_large.xlsx"

        # This should complete without memory errors
        try:
            translate_excel_with_format(input_file, output_file, "fr", "en")
            success = True
        except MemoryError:
            success = False

        assert success, "Memory error occurred with large file"

        # Cleanup
        if os.path.exists(output_file):
            os.remove(output_file)

    def test_multiple_large_operations(self):
        """Test handling multiple large operations sequentially"""
        input_file = "test_data/large_french.xlsx"

        # Process the same large file multiple times
        for i in range(3):
            output_file = f"test_results/memory_multi_{i}.xlsx"
            translate_excel_with_format(input_file, output_file, "fr", "en")

            # Cleanup immediately to free memory
            os.remove(output_file)

        # If we get here, no memory issues occurred
        assert True


class TestScalability:
    """Test scalability with increasing data sizes"""

    def test_scaling_by_rows(self):
        """Test performance scaling with increasing rows"""
        row_counts = [10, 50, 100, 500]
        times = []

        for row_count in row_counts:
            # Create test file
            wb = Workbook()
            ws = wb.active

            for row in range(1, row_count + 1):
                ws[f'A{row}'] = f"Ligne {row}"
                ws[f'B{row}'] = "Bonjour"
                ws[f'C{row}'] = "Merci"

            input_file = f"test_data/scale_rows_{row_count}.xlsx"
            output_file = f"test_results/scale_rows_{row_count}.xlsx"
            wb.save(input_file)

            # Time the translation
            start = time.time()
            translate_excel_with_format(input_file, output_file, "fr", "en")
            duration = time.time() - start

            times.append(duration)
            print(f"\n{row_count} rows: {duration:.2f}s")

            # Cleanup
            os.remove(input_file)
            os.remove(output_file)

        # Check that time increases roughly linearly (not exponentially)
        # This is a loose check - just ensure it's not wildly inefficient
        assert times[-1] / times[0] < (row_counts[-1] / row_counts[0]) * 2

    def test_scaling_by_columns(self):
        """Test performance scaling with increasing columns"""
        col_counts = [5, 10, 20, 50]
        times = []

        for col_count in col_counts:
            # Create test file
            wb = Workbook()
            ws = wb.active

            for col in range(1, col_count + 1):
                ws.cell(row=1, column=col, value=f"Colonne {col}")
                ws.cell(row=2, column=col, value="Bonjour")

            input_file = f"test_data/scale_cols_{col_count}.xlsx"
            output_file = f"test_results/scale_cols_{col_count}.xlsx"
            wb.save(input_file)

            # Time the translation
            start = time.time()
            translate_excel_with_format(input_file, output_file, "fr", "en")
            duration = time.time() - start

            times.append(duration)
            print(f"\n{col_count} columns: {duration:.2f}s")

            # Cleanup
            os.remove(input_file)
            os.remove(output_file)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
