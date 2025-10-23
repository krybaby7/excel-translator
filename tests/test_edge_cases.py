"""
Edge case and stress tests for Excel Translator
"""
import pytest
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from excel_translator import translate_excel_with_format, convert_xls_to_xlsx


class TestEdgeCases:
    """Test edge cases and unusual scenarios"""

    def test_empty_cells_mixed_with_text(self):
        """Test file with empty cells interspersed with text"""
        # Create test file
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Bonjour"
        ws['A2'] = None
        ws['A3'] = "Au revoir"
        ws['B1'] = None
        ws['B2'] = "Merci"

        test_file = "test_data/edge_empty_cells.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_empty_cells_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        # Verify
        wb_result = load_workbook(output_file)
        ws_result = wb_result.active
        assert ws_result['A1'].value is not None
        assert ws_result['A2'].value is None
        assert ws_result['A3'].value is not None

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_very_long_text_in_cell(self):
        """Test cell with very long text (>5000 characters)"""
        wb = Workbook()
        ws = wb.active

        long_text = "Bonjour " * 1000  # Very long repeated text
        ws['A1'] = long_text

        test_file = "test_data/edge_long_text.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_long_text_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        # Verify file was created
        assert os.path.exists(output_file)

        wb_result = load_workbook(output_file)
        ws_result = wb_result.active
        assert ws_result['A1'].value is not None

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_all_numeric_content(self):
        """Test file with only numbers (no text to translate)"""
        wb = Workbook()
        ws = wb.active

        for i in range(1, 11):
            ws[f'A{i}'] = i * 100

        test_file = "test_data/edge_all_numbers.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_all_numbers_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        # Verify numbers are unchanged
        wb_result = load_workbook(output_file)
        ws_result = wb_result.active

        for i in range(1, 11):
            assert ws_result[f'A{i}'].value == i * 100

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_formulas_preserved(self):
        """Test that Excel formulas are preserved"""
        wb = Workbook()
        ws = wb.active

        ws['A1'] = 10
        ws['A2'] = 20
        ws['A3'] = '=SUM(A1:A2)'
        ws['B1'] = "Total"

        test_file = "test_data/edge_formulas.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_formulas_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        wb_result = load_workbook(output_file)
        ws_result = wb_result.active

        # Check formula is still there
        assert ws_result['A3'].value == '=SUM(A1:A2)' or ws_result['A3'].value == 30

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_merged_cells(self):
        """Test handling of merged cells"""
        input_file = "test_data/merged_cells.xlsx"
        output_file = "test_results/edge_merged_cells_translated.xlsx"

        translate_excel_with_format(input_file, output_file, "fr", "en")

        # Verify merged cells still exist
        wb = load_workbook(output_file)
        ws = wb.active

        # Check if merged cells exist
        assert len(ws.merged_cells.ranges) > 0

        # Cleanup
        os.remove(output_file)

    def test_special_characters_preserved(self):
        """Test that special characters are properly handled"""
        input_file = "test_data/special_chars_french.xlsx"
        output_file = "test_results/edge_special_chars_translated.xlsx"

        translate_excel_with_format(input_file, output_file, "fr", "en")

        wb = load_workbook(output_file)
        ws = wb.active

        # Verify some content exists (emojis, symbols, etc.)
        assert ws['A2'].value is not None  # Emojis
        assert ws['A3'].value is not None  # Symbols

        # Cleanup
        os.remove(output_file)

    def test_rtl_languages(self):
        """Test right-to-left languages like Arabic"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "مرحبا"  # Arabic for "Hello"
        ws['A2'] = "كيف حالك"  # "How are you"

        test_file = "test_data/edge_arabic.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_arabic_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "ar", "en")

        assert os.path.exists(output_file)

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_mixed_languages_in_same_file(self):
        """Test file with multiple languages in different cells"""
        input_file = "test_data/various_languages.xlsx"
        output_file = "test_results/edge_mixed_languages_translated.xlsx"

        # This might produce unexpected results but shouldn't crash
        translate_excel_with_format(input_file, output_file, "auto", "en")

        assert os.path.exists(output_file)

        # Cleanup
        os.remove(output_file)

    def test_single_row_file(self):
        """Test file with only one row"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Bonjour"
        ws['B1'] = "Monde"
        ws['C1'] = "Merci"

        test_file = "test_data/edge_single_row.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_single_row_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        assert os.path.exists(output_file)

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_single_column_file(self):
        """Test file with only one column"""
        wb = Workbook()
        ws = wb.active

        french_words = ["Bonjour", "Merci", "Au revoir", "S'il vous plaît"]
        for i, word in enumerate(french_words, 1):
            ws[f'A{i}'] = word

        test_file = "test_data/edge_single_column.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_single_column_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        assert os.path.exists(output_file)

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_whitespace_only_cells(self):
        """Test cells containing only whitespace"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "   "
        ws['A2'] = "\t\t"
        ws['A3'] = "\n\n"
        ws['A4'] = "Bonjour"

        test_file = "test_data/edge_whitespace.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_whitespace_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        assert os.path.exists(output_file)

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_number_as_text(self):
        """Test numbers stored as text"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "123"  # String that looks like number
        ws['A2'] = 456    # Actual number
        ws['A3'] = "789.12"

        test_file = "test_data/edge_number_as_text.xlsx"
        wb.save(test_file)

        output_file = "test_results/edge_number_as_text_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        wb_result = load_workbook(output_file)
        ws_result = wb_result.active

        # Numbers should remain unchanged
        assert ws_result['A2'].value == 456

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)


class TestStressTesting:
    """Stress tests for performance and limits"""

    def test_very_wide_spreadsheet(self):
        """Test spreadsheet with many columns (100 columns)"""
        wb = Workbook()
        ws = wb.active

        for col in range(1, 101):
            ws.cell(row=1, column=col, value=f"Colonne {col}")

        test_file = "test_data/stress_wide.xlsx"
        wb.save(test_file)

        output_file = "test_results/stress_wide_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        assert os.path.exists(output_file)

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_many_sheets(self):
        """Test workbook with many sheets (20 sheets)"""
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        for i in range(1, 21):
            ws = wb.create_sheet(f"Feuille{i}")
            ws['A1'] = f"Contenu de la feuille {i}"

        test_file = "test_data/stress_many_sheets.xlsx"
        wb.save(test_file)

        output_file = "test_results/stress_many_sheets_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        # Verify all sheets were processed
        wb_result = load_workbook(output_file)
        assert len(wb_result.sheetnames) == 20

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)

    def test_complex_formatting(self):
        """Test file with complex formatting combinations"""
        wb = Workbook()
        ws = wb.active

        # Cell with multiple formatting attributes
        ws['A1'] = "Texte complexe"
        ws['A1'].font = Font(bold=True, italic=True, size=14, color="FF0000", underline="single")
        ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        test_file = "test_data/stress_complex_format.xlsx"
        wb.save(test_file)

        output_file = "test_results/stress_complex_format_translated.xlsx"
        translate_excel_with_format(test_file, output_file, "fr", "en")

        # Verify formatting preserved
        wb_result = load_workbook(output_file)
        ws_result = wb_result.active

        assert ws_result['A1'].font.bold
        assert ws_result['A1'].font.italic

        # Cleanup
        os.remove(test_file)
        os.remove(output_file)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
