"""
Excel Translator Module
Translates Excel files while preserving formatting
"""
import os
import logging
import re
from copy import copy
from openpyxl import load_workbook, Workbook
from deep_translator import GoogleTranslator
import xlrd

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


def convert_xls_to_xlsx(xls_file):
    """Convert .xls file to .xlsx using xlrd and openpyxl."""
    if not os.path.exists(xls_file):
        raise FileNotFoundError(f"File not found: {xls_file}")

    if not xls_file.endswith('.xls'):
        raise ValueError("Input file must be .xls format")

    wb_xls = xlrd.open_workbook(xls_file)
    wb_xlsx = Workbook()
    sheet_names = wb_xls.sheet_names()

    for sheet_index, sheet_name in enumerate(sheet_names):
        xls_sheet = wb_xls.sheet_by_index(sheet_index)

        # Remove default sheet on first iteration
        if sheet_index == 0:
            default_sheet = wb_xlsx.active
            wb_xlsx.remove(default_sheet)

        xlsx_sheet = wb_xlsx.create_sheet(title=sheet_name)

        # Copy content from .xls to .xlsx sheet
        for row_index in range(xls_sheet.nrows):
            for col_index in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row_index, col_index)
                xlsx_sheet.cell(row=row_index + 1, column=col_index + 1, value=cell_value)

    xlsx_file = xls_file.replace(".xls", ".xlsx")
    wb_xlsx.save(xlsx_file)
    return xlsx_file


def should_translate_string(text, formula):
    """
    Determine if a string literal should be translated based on context.

    Args:
        text: The string content (without quotes)
        formula: The full formula containing this string

    Returns:
        True if the string should be translated, False if it should be preserved
    """
    # Skip empty strings
    if not text.strip():
        return False

    # Skip if formula contains known technical/data functions where parameters shouldn't be translated
    technical_functions = ['SPARKLINE', '__xludf.DUMMYFUNCTION', 'IMPORTDATA', 'QUERY', 'GOOGLETRANSLATE']
    if any(func in formula.upper() for func in technical_functions):
        return False

    # Skip technical-looking strings (single lowercase words without spaces)
    # These are typically parameter names like "charttype", "column", "max"
    if len(text.split()) == 1 and text.islower() and text.isalpha():
        return False

    # Skip hex color codes
    if re.match(r'^#[0-9A-Fa-f]{6}$', text):
        return False

    # Skip very short strings (1-2 chars) that are likely technical
    if len(text) <= 2:
        return False

    # Translate everything else (user-facing text)
    return True


def translate_formula_strings(formula, translator):
    """
    Translate string literals inside Excel formulas while preserving formula structure.

    Example:
        Input: =if(J15<0, "Spent this month", "Saved this month")
        Output: =if(J15<0, "Dépensé ce mois-ci", "Économisé ce mois-ci")

    Args:
        formula: Excel formula string starting with =
        translator: GoogleTranslator instance

    Returns:
        Formula with translated string literals
    """
    if not formula.startswith('='):
        return formula

    # Find all string literals in the formula (text in double quotes)
    # Pattern matches: "any text except quotes"
    # This handles most cases but not escaped quotes
    pattern = r'"([^"]*)"'

    def translate_match(match):
        original_text = match.group(1)

        # Use smart filtering to determine if this string should be translated
        if not should_translate_string(original_text, formula):
            return match.group(0)

        try:
            translated_text = translator.translate(original_text)
            return f'"{translated_text}"'
        except Exception as e:
            # If translation fails, keep original
            logger.warning(f"Failed to translate formula string '{original_text[:30]}...': {e}")
            return match.group(0)

    # Replace all quoted strings with their translations
    translated_formula = re.sub(pattern, translate_match, formula)
    return translated_formula


def translate_excel_with_format(input_file, output_file, source_lang="fr", target_lang="en", progress_callback=None):
    """Translate text in an Excel file, preserving formatting.

    Args:
        input_file: Path to input Excel file
        output_file: Path to save translated file
        source_lang: Source language code
        target_lang: Target language code
        progress_callback: Optional callback function(current, total, message) for progress updates
    """
    # Check format FIRST before checking file existence
    if not input_file.endswith('.xlsx'):
        raise ValueError("Input file must be .xlsx format")

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"File not found: {input_file}")

    logger.info(f"Starting translation: {input_file}")
    logger.info(f"Languages: {source_lang} -> {target_lang}")

    if progress_callback:
        progress_callback(0, 0, f"Starting translation: {source_lang} -> {target_lang}")

    # Load workbook
    wb = load_workbook(input_file)
    translator = GoogleTranslator(source=source_lang, target=target_lang)

    total_sheets = len(wb.sheetnames)
    logger.info(f"Found {total_sheets} sheet(s) to process")

    # Count total text cells and formulas separately
    # We'll translate both, but differently (text fully, formulas only their string literals)
    total_text_cells = 0
    total_formula_cells = 0
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith('='):
                        total_formula_cells += 1
                    else:
                        total_text_cells += 1

    total_cells = total_text_cells + total_formula_cells
    logger.info(f"Found {total_text_cells} text cells and {total_formula_cells} formulas to translate")

    if progress_callback:
        progress_callback(0, total_cells, f"Found {total_cells} cells to translate ({total_text_cells} text + {total_formula_cells} formulas)")

    global_cell_count = 0

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        logger.info(f"Processing sheet {sheet_idx}/{total_sheets}: '{sheet_name}'")

        # Count text cells and formulas separately for this sheet
        sheet_text_cells = 0
        sheet_formulas = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith('='):
                        sheet_formulas += 1
                    else:
                        sheet_text_cells += 1

        sheet_total = sheet_text_cells + sheet_formulas
        logger.info(f"Sheet '{sheet_name}': {sheet_text_cells} text cells, {sheet_formulas} formulas")

        # Iterate through rows and columns
        translated_count = 0
        formula_count = 0
        error_count = 0

        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):  # Check if cell contains text
                    global_cell_count += 1

                    try:
                        # Preserve cell formatting
                        old_alignment = copy(cell.alignment)
                        old_font = copy(cell.font)
                        old_fill = copy(cell.fill)
                        old_border = copy(cell.border)
                        old_number_format = cell.number_format

                        if cell.value.startswith('='):
                            # This is a formula - translate only string literals inside it
                            original_formula = cell.value
                            translated_formula = translate_formula_strings(cell.value, translator)
                            cell.value = translated_formula
                            formula_count += 1

                            # Log if formula strings were translated
                            if original_formula != translated_formula:
                                logger.debug(f"Formula translated at {cell.coordinate}")

                        else:
                            # This is plain text - translate it fully
                            original_value = cell.value
                            translated = translator.translate(cell.value)
                            cell.value = translated
                            translated_count += 1

                        # Restore cell formatting
                        cell.alignment = old_alignment
                        cell.font = old_font
                        cell.fill = old_fill
                        cell.border = old_border
                        cell.number_format = old_number_format

                        # Send progress update
                        if progress_callback:
                            progress_pct = int(global_cell_count/total_cells*100) if total_cells > 0 else 0
                            progress_callback(global_cell_count, total_cells,
                                            f"Translating '{sheet_name}': {global_cell_count}/{total_cells} cells ({progress_pct}%)")

                        # Log progress every 10 cells
                        if (translated_count + formula_count) % 10 == 0:
                            progress_pct = int((translated_count + formula_count)/sheet_total*100) if sheet_total > 0 else 0
                            logger.info(f"Sheet '{sheet_name}': {translated_count + formula_count}/{sheet_total} cells processed ({progress_pct}%)")

                    except Exception as e:
                        # Keep original value if translation fails
                        error_count += 1
                        cell_preview = str(cell.value)[:50] if cell.value else ""
                        logger.warning(f"Translation failed for cell ({row_idx},{col_idx}): '{cell_preview}...' - Error: {e}")
                        pass

        logger.info(f"Sheet '{sheet_name}' complete: {translated_count} text cells, {formula_count} formulas processed, {error_count} errors")

    logger.info(f"Saving translated workbook to: {output_file}")
    if progress_callback:
        progress_callback(total_cells, total_cells, "Saving translated file...")

    # Save the translated workbook
    wb.save(output_file)
    logger.info(f"Translation complete! File saved: {output_file}")
    logger.info(f"Summary: {total_text_cells} text cells translated, {total_formula_cells} formulas processed")

    if progress_callback:
        progress_callback(total_cells, total_cells, "Translation complete!")

    return output_file


def process_file(input_filename, source_lang="fr", target_lang="en"):
    """Process a single file - convert if needed and translate."""
    original_file = input_filename

    # Convert .xls to .xlsx if needed
    if input_filename.endswith(".xls"):
        input_filename = convert_xls_to_xlsx(input_filename)

    if input_filename.endswith(".xlsx"):
        output_filename = f"translated_{os.path.basename(input_filename)}"
        translate_excel_with_format(input_filename, output_filename, source_lang, target_lang)
        return output_filename
    else:
        raise ValueError(f"Unsupported file format: {input_filename}")
