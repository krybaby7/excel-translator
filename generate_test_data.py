"""
Generate test data for Excel Translator testing
Creates various Excel files with different scenarios
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import xlwt
import os


def create_simple_french_file():
    """Create a simple French Excel file for basic testing"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Simple"

    # Add some French text
    ws['A1'] = "Bonjour"
    ws['A2'] = "Comment allez-vous?"
    ws['A3'] = "Merci beaucoup"
    ws['B1'] = "Monde"
    ws['B2'] = "Bien, merci"

    wb.save('test_data/simple_french.xlsx')
    print("Created: test_data/simple_french.xlsx")


def create_formatted_file():
    """Create Excel file with various formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formatted"

    # Add text with different formatting
    ws['A1'] = "Titre Important"
    ws['A1'].font = Font(bold=True, size=16, color="FF0000")
    ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws['A2'] = "Texte en italique"
    ws['A2'].font = Font(italic=True)

    ws['A3'] = "Texte souligné"
    ws['A3'].font = Font(underline="single")

    ws['B1'] = "Centré"
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center")

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws['C1'] = "Avec bordure"
    ws['C1'].border = thin_border

    wb.save('test_data/formatted_french.xlsx')
    print("Created: test_data/formatted_french.xlsx")


def create_multi_sheet_file():
    """Create Excel file with multiple sheets"""
    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Feuille1"
    ws1['A1'] = "Première feuille"
    ws1['A2'] = "Quelques données"

    # Sheet 2
    ws2 = wb.create_sheet("Feuille2")
    ws2['A1'] = "Deuxième feuille"
    ws2['A2'] = "Plus de données"

    # Sheet 3
    ws3 = wb.create_sheet("Feuille3")
    ws3['A1'] = "Troisième feuille"
    ws3['A2'] = "Encore plus de données"

    wb.save('test_data/multi_sheet_french.xlsx')
    print("Created: test_data/multi_sheet_french.xlsx")


def create_mixed_content_file():
    """Create file with mixed content types"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"

    ws['A1'] = "Nom"
    ws['B1'] = "Âge"
    ws['C1'] = "Date"

    ws['A2'] = "Jean"
    ws['B2'] = 25
    ws['C2'] = "2024-01-15"

    ws['A3'] = "Marie"
    ws['B3'] = 30
    ws['C3'] = "2024-02-20"

    # Add a formula
    ws['A4'] = "Total"
    ws['B4'] = "=SUM(B2:B3)"

    # Empty cell
    ws['A5'] = None

    # Number stored as text
    ws['A6'] = "123"

    wb.save('test_data/mixed_content_french.xlsx')
    print("Created: test_data/mixed_content_french.xlsx")


def create_special_characters_file():
    """Create file with special characters"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Special"

    ws['A1'] = "Caractères spéciaux: é, è, ê, à, ù, ç"
    ws['A2'] = "Emojis: 😀 🎉 ❤️"
    ws['A3'] = "Symboles: © ® ™ € £ ¥"
    ws['A4'] = "Citations: \"Bonjour\" et 'Au revoir'"
    ws['A5'] = "Newline\ndans\ncellule"
    ws['A6'] = "Tab\tentre\tmots"

    wb.save('test_data/special_chars_french.xlsx')
    print("Created: test_data/special_chars_french.xlsx")


def create_large_file():
    """Create a larger file for performance testing"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Large"

    french_phrases = [
        "Bonjour", "Comment allez-vous?", "Merci beaucoup",
        "S'il vous plaît", "Au revoir", "Bonne journée",
        "Excusez-moi", "Je ne comprends pas", "Pouvez-vous m'aider?",
        "C'est délicieux"
    ]

    # Create 1000 rows with 10 columns
    for row in range(1, 1001):
        for col in range(1, 11):
            phrase_index = (row + col) % len(french_phrases)
            ws.cell(row=row, column=col, value=french_phrases[phrase_index])

    wb.save('test_data/large_french.xlsx')
    print("Created: test_data/large_french.xlsx (1000x10 cells)")


def create_empty_file():
    """Create an empty Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"

    wb.save('test_data/empty.xlsx')
    print("Created: test_data/empty.xlsx")


def create_single_cell_file():
    """Create file with single cell"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bonjour"

    wb.save('test_data/single_cell.xlsx')
    print("Created: test_data/single_cell.xlsx")


def create_xls_file():
    """Create an old format .xls file"""
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')

    ws.write(0, 0, 'Bonjour')
    ws.write(0, 1, 'Monde')
    ws.write(1, 0, 'Comment')
    ws.write(1, 1, 'allez-vous?')

    wb.save('test_data/old_format.xls')
    print("Created: test_data/old_format.xls")


def create_merged_cells_file():
    """Create file with merged cells"""
    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A1:C1')
    ws['A1'] = "Titre fusionné"

    ws['A2'] = "Cellule normale"
    ws['B2'] = "Autre cellule"

    wb.save('test_data/merged_cells.xlsx')
    print("Created: test_data/merged_cells.xlsx")


def create_various_languages_file():
    """Create file with text in various languages for testing"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Languages"

    ws['A1'] = "Bonjour"  # French
    ws['B1'] = "Français"

    ws['A2'] = "Hola"  # Spanish
    ws['B2'] = "Espagnol"

    ws['A3'] = "Ciao"  # Italian
    ws['B3'] = "Italien"

    ws['A4'] = "Hallo"  # German
    ws['B4'] = "Allemand"

    ws['A5'] = "こんにちは"  # Japanese
    ws['B5'] = "Japonais"

    ws['A6'] = "مرحبا"  # Arabic
    ws['B6'] = "Arabe"

    wb.save('test_data/various_languages.xlsx')
    print("Created: test_data/various_languages.xlsx")


if __name__ == "__main__":
    # Create test_data directory if it doesn't exist
    os.makedirs('test_data', exist_ok=True)

    print("Generating test data files...")
    print("-" * 50)

    create_simple_french_file()
    create_formatted_file()
    create_multi_sheet_file()
    create_mixed_content_file()
    create_special_characters_file()
    create_large_file()
    create_empty_file()
    create_single_cell_file()
    create_xls_file()
    create_merged_cells_file()
    create_various_languages_file()

    print("-" * 50)
    print("Test data generation complete!")
    print(f"Files created in: {os.path.abspath('test_data')}")
